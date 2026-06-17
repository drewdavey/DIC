#!/usr/bin/env python3
"""
DIC_Level3.py  —  FSR Tensile Coupons
======================================
Reads Level-2 result CSVs, computes mechanical properties per
ASTM D638-14, and saves per-coupon σ-ε plots plus group overlays
and a property summary scatter to FIGS_ROOT.

Standards compliance — what each calculation cites
  Toe compensation     : D638 Annex A1 (mandatory unless toe is real material response)
  Modulus              : D638 §11.4   (slope of initial linear region of σ-ε)
  0.2 % offset yield   : D638 §A2.6 / Fig. A2.1 (offset from toe-corrected origin)
  Tensile strength UTS : D638 §11.2   (max stress / original area)
  Poisson's ratio      : D638 Annex A3.10.1.3 (chord at ε_a=0.002 over 0.0005-0.0025)
  Group statistics     : D638 §11.7 / §12.1   (mean, std per series)

PLOTTING NOTE
  Level-2 writes the full, untruncated DIC record (see its docstring). All
  failure truncation happens here via truncate_df(): pre-load slack and
  post-fracture rebound (past 50% post-UTS load drop) are cut, then
  per-coupon plots are truncated again at exactly UTS so post-fracture
  rebound (which appears as the curve doubling back) is never shown, since
  post-UTS data violates the monotonic-loading assumption D638 calculations
  rely on.

OUTPUT
  {FIGS_ROOT}/{coupon_id}/stress_strain_DIC.png   per-coupon σ–ε (toe-corrected, to UTS)
  {FIGS_ROOT}/{coupon_id}/poisson_DIC.png         per-coupon −ε_xx vs ε_yy (to UTS)
  {FIGS_ROOT}/tensile_curves_DIC.png              group overlay (1 panel per direction)
  {FIGS_ROOT}/tensile_summary_DIC.png             property scatter (E, σ_y, UTS)
  {FIGS_ROOT}/level3_summary.csv                  per-coupon property table
  {FIGS_ROOT}/level3_group_stats.csv              D638 §11.7 mean/std by exposure×direction
"""

from __future__ import annotations
import time
from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import matplotlib.patches as mpatches
from scipy.ndimage import median_filter
import openpyxl

# =============================================================================
# PATHS
# =============================================================================
DATA_ROOTS = {
    "CL": Path(r"G:\DrewDavey\2026_FSR_TensileTest_TCL"),
    "SW": Path(r"G:\DrewDavey\2026_FSR_TensileTest_TSW_TIS_TUV"),
    "UV": Path(r"G:\DrewDavey\2026_FSR_TensileTest_TSW_TIS_TUV"),
    "IS": Path(r"G:\DrewDavey\2026_FSR_TensileTest_TSW_TIS_TUV"),
}
FIGS_ROOT = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\04_TestCoupons"
    r"\P01-LT150-LH4.5\figs"
)
DIC_DIR = FIGS_ROOT.parent / "DIC"   # consolidated _L2.csv files written by Level 2
SPECIMEN_SHEET = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\FSR-SpecimenTesting.xlsx"
)

# =============================================================================
# SWITCHES
# =============================================================================
PRINTS     = ["P01"]
EXPOSURES  = {"CL": True, "UV": True, "SW": True, "IS": True}
# DIRECTIONS = {"00": True, "45": True, "90": True}
DIRECTIONS = {"00": True, "45": False, "90": False}
# REPLICATES = ["01", "02", "03"]
REPLICATES = ["01"]

# =============================================================================
# FAILURE TRUNCATION  — applied to Level-2 data before property extraction
# Trim pre-load slack and post-fracture rebound so only the valid test window
# is passed to compute_properties and the plots.
# =============================================================================
LOAD_START_FRAC = 0.02     # pre-load: drop frames before load exceeds this × peak
LOAD_END_FRAC   = 0.50     # post-fracture: cut first post-UTS frame where load < this × peak

# Scale factor applied to load_raw to produce force_N (N per sync-CSV unit).
# Two modes — the first one found is used:
#   1. Per-coupon  : mts_peak_N / max(raw load_raw)  — most accurate
#   2. Combined    : SCALE_N_PER_UNIT below          — fallback if mts_peak_N missing
# Set SCALE_N_PER_UNIT to the mean reported by Level-2's calibration pass.
SCALE_N_PER_UNIT: float = 555.5928

# =============================================================================
# PROPERTY SETTINGS
# =============================================================================
# Modulus fit window (axial strain, dimensionless).
# D638 §11.4: "initial linear portion of the load-extension curve".
# A window of 0.05–0.3% covers the typical linear region for stiff polymers
# and composites without including the toe. Adjust if the fit line on the
# generated plot doesn't sit on the linear segment.
MODULUS_STRAIN_RANGE = (0.0005, 0.003)

# D638 §A2.6 — 0.2% offset yield strength
YIELD_OFFSET = 0.002

# D638 §A3.10.1.3 — Poisson chord method window (when no clear proportionality)
# Chord computed at ε_a = 0.002 over the range 0.0005 to 0.0025 strain.
POISSON_RANGE = (0.0005, 0.0025)
POISSON_CHORD_AT = 0.002

# Airtech reference values (printed material spec — comparison lines)
AIRTECH_UTS = {0: 79.3, 45: None, 90: 25.9}    # MPa
AIRTECH_E   = {0: 6.6,  45: None, 90: 3.7}     # GPa

# Scalar property columns written into SPECIMEN_SHEET, keyed by coupon
# ("Specimen ID") — maps the property dict key to the Excel column header.
SPECIMEN_SHEET_COLUMNS = {
    "E_GPa":         "E (GPa)",
    "eps_toe":       "Toe Strain",
    "sigma_y_MPa":   "Yield Stress (MPa)",
    "eps_y":         "Yield Strain",
    "UTS_MPa":       "UTS (MPa)",
    "eps_at_UTS":    "Strain at UTS",
    "poisson_chord": "Poisson's Ratio (chord)",
    "poisson_slope": "Poisson's Ratio (slope)",
}

# =============================================================================
# DISPLAY
# =============================================================================
EXPOSURE_COLORS = {"CL": "#1f77b4", "SW": "#17becf", "UV": "#ff7f0e", "IS": "#2ca02c"}
EXPOSURE_LABELS = {"CL": "Control", "UV": "UV", "SW": "Seawater", "IS": "SW+UV"}
DIRECTION_MARKERS = {"00": "o", "45": "s", "90": "^"}

# =============================================================================
# HELPERS
# =============================================================================
def coupon_id(p, e, d, r): return f"{p}-T{e}{d}-{r}"

def selected_coupons():
    return [coupon_id(p, e, d, r)
            for p in PRINTS
            for e, on in EXPOSURES.items() if on
            for d, on2 in DIRECTIONS.items() if on2
            for r in REPLICATES]

def parse_id(cid):
    """Return (exposure, direction_str) e.g. ('CL', '00')."""
    part = cid.split("-")[1]
    return part[1:-2], part[-2:]

def find_l2(cid):
    p = DIC_DIR / f"{cid}_L2.csv"
    return p if p.exists() else None

def write_specimen_sheet(rows: list[dict]) -> None:
    """Write each coupon's scalar properties into its row in SPECIMEN_SHEET,
    matched by Specimen ID. Adds any missing property columns at the end;
    everything else in the workbook (other rows, formulas, formatting) is
    left untouched. Skipped (with a warning) if the file can't be opened —
    e.g. if it's currently open in Excel.
    """
    try:
        wb = openpyxl.load_workbook(SPECIMEN_SHEET)
    except FileNotFoundError:
        print(f"[!] {SPECIMEN_SHEET} not found — skipping specimen sheet update")
        return
    ws = wb.active

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    id_col = header.get("Specimen ID")
    if id_col is None:
        print("[!] 'Specimen ID' column not found in specimen sheet — skipping update")
        return

    next_col = ws.max_column + 1
    for label in SPECIMEN_SHEET_COLUMNS.values():
        if label not in header:
            ws.cell(row=1, column=next_col, value=label)
            header[label] = next_col
            next_col += 1

    row_by_id = {ws.cell(row=r, column=id_col).value: r
                 for r in range(2, ws.max_row + 1)}

    for row in rows:
        r = row_by_id.get(row["coupon"])
        if r is None:
            continue
        for key, label in SPECIMEN_SHEET_COLUMNS.items():
            v = row.get(key)
            v = None if (v is None or not np.isfinite(v)) else v
            ws.cell(row=r, column=header[label], value=v)

    try:
        wb.save(SPECIMEN_SHEET)
    except PermissionError:
        print(f"[!] {SPECIMEN_SHEET} is open elsewhere — could not save properties to it")

def truncate_df(df: pd.DataFrame) -> pd.DataFrame:
    """Remove pre-load slack and post-fracture frames (Level-3 failure truncation)."""
    f = df["force_N"].to_numpy()
    peak = float(np.nanmax(np.abs(f)))
    if peak <= 0:
        return df
    i_uts = int(np.nanargmax(np.abs(f)))
    starts = np.where(np.abs(f) > LOAD_START_FRAC * peak)[0]
    i0 = int(starts[0]) if len(starts) else 0
    post = np.where(np.abs(f[i_uts:]) < LOAD_END_FRAC * peak)[0]
    i1 = int(i_uts + post[0]) if len(post) else len(f) - 1
    return df.iloc[i0:i1 + 1].reset_index(drop=True)


# =============================================================================
# SMOOTHING  (rolling median)
# Window must be odd. Raise MEDIAN_WINDOW for noisier data.
# =============================================================================
MEDIAN_WINDOW = 11  # frames

def smooth_signal(x):
    """Rolling median. mode='nearest' avoids the zero-padding edge artifacts
    scipy.signal.medfilt has."""
    x = np.asarray(x, dtype=float)
    win = MEDIAN_WINDOW
    if win % 2 == 0:
        win -= 1
    if win < 1 or len(x) < win:
        return x.copy()
    return median_filter(x, size=win, mode="constant", cval=np.nan)

# =============================================================================
# COMPUTE PROPERTIES
# =============================================================================
def compute_properties(df):
    """
    D638-compliant property extraction.

    Returns a dict with E_GPa, eps_toe (toe-correction offset applied),
    sigma_y_MPa, eps_y, UTS_MPa, eps_at_UTS, poisson_chord, poisson_slope,
    plus i_uts (index of UTS in the original arrays — used by plotter).
    Also returns the toe-corrected eps and (truncated-at-UTS) eps_t arrays
    so plotting uses the same numbers the properties were computed from.
    """
    df = df.dropna(subset=["strain_axial", "stress_MPa"]).reset_index(drop=True)
    if len(df) < 10:
        return None

    eps_raw   = df["strain_axial"].to_numpy()
    sig       = df["stress_MPa"].to_numpy()
    eps_t_raw = (df["strain_transverse"].to_numpy()
                 if "strain_transverse" in df.columns
                 else np.full_like(eps_raw, np.nan))
    # Pre-smoothing reference, carried through only for the diagnostic
    # overlay in plot_stress_strain() — not used in any property calculation.
    eps_unsmoothed = df["strain_axial_raw"].to_numpy() if "strain_axial_raw" in df.columns else None
    sig_unsmoothed = df["stress_MPa_raw"].to_numpy()   if "stress_MPa_raw"   in df.columns else None

    # ---- 1. Modulus (D638 §11.4) --------------------------------------------
    lo, hi = MODULUS_STRAIN_RANGE
    mfit = (eps_raw >= lo) & (eps_raw <= hi) & np.isfinite(eps_raw) & np.isfinite(sig)
    if mfit.sum() < 3:
        return None
    slope, intercept = np.polyfit(eps_raw[mfit], sig[mfit], 1)
    E_MPa = float(slope)

    # ---- 2. Toe compensation (D638 Annex A1) --------------------------------
    # The fitted line σ = E·ε + b is extended back to σ = 0; that strain
    # (b/(-E)) is the "toe offset" — all strains are then measured from the
    # corrected origin. ε_corrected = ε_raw − ε_offset.
    eps_offset = -intercept / E_MPa if E_MPa != 0 else 0.0
    eps   = eps_raw   - eps_offset
    eps_unsmoothed_corr = (eps_unsmoothed - eps_offset) if eps_unsmoothed is not None else None
    # Transverse strain: subtract its value at the corrected zero of axial strain.
    # Find the index where corrected axial ≈ 0 and subtract that ε_t.
    if np.any(np.isfinite(eps_t_raw)):
        i0 = int(np.nanargmin(np.abs(eps)))
        eps_t = eps_t_raw - eps_t_raw[i0]
    else:
        eps_t = eps_t_raw.copy()

    # ---- 3. UTS (D638 §11.2 — max stress) -----------------------------------
    i_uts   = int(np.nanargmax(sig))
    uts     = float(sig[i_uts])
    eps_ult = float(eps[i_uts])

    # ---- 4. 0.2% offset yield (D638 §A2.6, Fig. A2.1) -----------------------
    # First crossing of σ-ε curve with the line σ = E·(ε − YIELD_OFFSET).
    sigma_y, eps_y = np.nan, np.nan
    diff  = sig - E_MPa * (eps - YIELD_OFFSET)
    valid = np.where(eps > YIELD_OFFSET)[0]
    if len(valid) > 1:
        d = diff[valid]
        crossings = np.where(np.diff(np.sign(d)) < 0)[0]
        if len(crossings):
            k = valid[crossings[0]]
            denom = diff[k] - diff[k+1]
            f = diff[k] / denom if denom != 0 else 0.0
            eps_y   = float(eps[k] + f * (eps[k+1] - eps[k]))
            sigma_y = float(sig[k] + f * (sig[k+1] - sig[k]))

    # ---- 5. Poisson's ratio (D638 §A3.10.1.3, chord at ε_a = 0.002) --------
    #   ν = − ε_t(at ε_a = 0.002) / 0.002
    # Range 0.0005 – 0.0025 strain. Also report least-squares slope (§A3.10.1.1)
    # for transparency when proportionality holds.
    nu_chord = nu_slope = np.nan
    pm = ((eps >= POISSON_RANGE[0]) & (eps <= POISSON_RANGE[1]) &
          np.isfinite(eps) & np.isfinite(eps_t))
    if pm.sum() >= 3:
        nu_slope = float(-np.polyfit(eps[pm], eps_t[pm], 1)[0])
        order = np.argsort(eps[pm])
        ea, et = eps[pm][order], eps_t[pm][order]
        if ea[0] <= POISSON_CHORD_AT <= ea[-1]:
            nu_chord = float(-np.interp(POISSON_CHORD_AT, ea, et) / POISSON_CHORD_AT)

    return {
        "E_GPa":         E_MPa / 1000.0,
        "eps_toe":       eps_offset,
        "sigma_y_MPa":   sigma_y,
        "eps_y":         eps_y,
        "UTS_MPa":       uts,
        "eps_at_UTS":    eps_ult,
        "poisson_chord": nu_chord,
        "poisson_slope": nu_slope,
        "_eps":          eps,        # toe-corrected strain, used in property calcs
        "_sig":          sig,
        "_eps_t":        eps_t,
        "_i_uts":        i_uts,
        "_eps_raw":      eps_unsmoothed_corr,  # diagnostic overlay only
        "_sig_raw":      sig_unsmoothed,
    }


# =============================================================================
# PER-COUPON PLOTS
# =============================================================================
def plot_stress_strain(cid, props, fig_dir):
    """σ-ε curve, toe-corrected, truncated at UTS. The raw (pre-smoothing,
    but still truncated) signal is drawn behind it in light gray, with its
    own peak marked, for comparison."""
    exp, d_str = parse_id(cid)
    direction  = int(d_str)

    sl  = slice(0, props["_i_uts"] + 1)
    e_p = props["_eps"][sl] * 100   # % strain
    s_p = props["_sig"][sl]

    fig, ax = plt.subplots(figsize=(7, 4.8))

    eps_r, sig_r = props.get("_eps_raw"), props.get("_sig_raw")
    if eps_r is not None and sig_r is not None and np.any(np.isfinite(sig_r)):
        ax.plot(eps_r * 100, sig_r, lw=0.8, color="0.8", zorder=1,
                label="raw (unsmoothed)")
        i_raw = int(np.nanargmax(sig_r))
        ax.plot(eps_r[i_raw] * 100, sig_r[i_raw], "^", color="0.6", ms=8,
                zorder=2, label=f"raw UTS = {sig_r[i_raw]:.1f} MPa")

    ax.plot(e_p, s_p, lw=1.4, color=EXPOSURE_COLORS.get(exp, "#333"),
            label=cid, zorder=3)

    E_MPa = props["E_GPa"] * 1000.0
    if np.isfinite(E_MPa):
        # Elastic line through toe-corrected origin (D638 Annex A1)
        x_e = np.array([0.0, MODULUS_STRAIN_RANGE[1] * 1.5])
        ax.plot(x_e * 100, E_MPa * x_e, "k--", lw=0.8, alpha=0.7,
                label=f"E = {props['E_GPa']:.1f} GPa")
        # 0.2% offset line — start at YIELD_OFFSET on the toe-corrected axis
        x_o_end = max(YIELD_OFFSET + 0.005,
                      props["eps_y"] if np.isfinite(props["eps_y"]) else YIELD_OFFSET + 0.005)
        x_o = np.linspace(YIELD_OFFSET, x_o_end, 50)
        ax.plot(x_o * 100, E_MPa * (x_o - YIELD_OFFSET), "k:", lw=0.8, alpha=0.6,
                label="0.2% offset")
    if np.isfinite(props["sigma_y_MPa"]):
        ax.plot(props["eps_y"] * 100, props["sigma_y_MPa"], "o",
                color="orange", ms=7, zorder=5,
                label=f"σ_y = {props['sigma_y_MPa']:.1f} MPa")
    ax.plot(props["eps_at_UTS"] * 100, props["UTS_MPa"], "^",
            color="red", ms=8, zorder=5,
            label=f"UTS = {props['UTS_MPa']:.1f} MPa")

    ref_uts = AIRTECH_UTS.get(direction)
    if ref_uts is not None:
        ax.axhline(ref_uts, color="grey", linestyle=":", lw=1.0, alpha=0.7,
                   label=f"Airtech UTS = {ref_uts} MPa")

    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    ax.set_xlabel("Axial Strain (%)")
    ax.set_ylabel("Engineering Stress (MPa)")
    ax.set_title(f"{cid}  —  {EXPOSURE_LABELS.get(exp, exp)}, {direction}°")
    ax.grid(alpha=0.25, linestyle="--")
    ax.legend(fontsize=8, framealpha=0.85, loc="best")
    fig.tight_layout()
    out = fig_dir / "stress_strain_DIC.png"
    fig.savefig(out, dpi=600, bbox_inches="tight")
    plt.close(fig)
    return out

def plot_poisson(cid, props, fig_dir):
    """−ε_xx vs ε_yy, truncated at UTS."""
    eps   = props["_eps"]
    eps_t = props["_eps_t"]
    if not np.any(np.isfinite(eps_t)):
        return None
    i_uts = props["_i_uts"]
    sl = slice(0, i_uts + 1)
    e_p, et_p = eps[sl], eps_t[sl]

    fig, ax = plt.subplots(figsize=(6, 4.5))
    ax.plot(e_p * 100, -et_p * 100, lw=1.2, label="data")
    nu_c = props["poisson_chord"]
    nu_s = props["poisson_slope"]
    if np.isfinite(nu_s):
        # show fit line over Poisson range
        x = np.linspace(POISSON_RANGE[0], POISSON_RANGE[1], 20)
        ax.plot(x * 100, nu_s * x * 100, "k--", lw=0.8, alpha=0.7,
                label=f"slope ν = {nu_s:.3f}")
    if np.isfinite(nu_c):
        ax.axvline(POISSON_CHORD_AT * 100, color="orange", ls=":", lw=0.8, alpha=0.6)
        # ax.set_title(f"{cid}     ν_chord = {nu_c:.3f}     "
        #              f"(D638 §A3.10.1.3, chord at ε_a = {POISSON_CHORD_AT*100:.1f}%)")
    else:
        ax.set_title(cid)
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    ax.set_xlabel("Axial strain ε_yy (%)")
    ax.set_ylabel("−Transverse strain  −ε_xx (%)")
    ax.grid(alpha=0.25, linestyle="--")
    ax.legend(fontsize=8)
    fig.tight_layout()
    out = fig_dir / "poisson_DIC.png"
    fig.savefig(out, dpi=600, bbox_inches="tight")
    plt.close(fig)
    return out

# =============================================================================
# MAIN
# =============================================================================
def main():
    t0 = time.time()
    FIGS_ROOT.mkdir(parents=True, exist_ok=True)
    props_by_cid = {}
    rows = []

    for cid in selected_coupons():
        l2 = find_l2(cid)
        if l2 is None:
            print(f"[{cid}] no _L2.csv — run Level 2 first")
            continue
        df = pd.read_csv(l2)

        if "load_raw" in df.columns:
            # Current L2 format: scale from raw peaks (mirrors Level-2's own
            # calibration pass).
            raw_peak = float(np.nanmax(np.abs(df["load_raw"].to_numpy())))
            if "mts_peak_N" in df.columns and raw_peak > 0:
                # Per-coupon scale: each coupon's own MTS peak / its own raw sync peak
                mts_peak = float(df["mts_peak_N"].iloc[0])
                scale = mts_peak / raw_peak
                print(f"[{cid}] per-coupon scale: {scale:.4f} N/unit  "
                      f"(MTS {mts_peak:.0f} N)")
            else:
                scale = SCALE_N_PER_UNIT
                print(f"[{cid}] combined scale: {scale:.4f} N/unit")
            area = float(df["area_mm2"].iloc[0]) if "area_mm2" in df.columns else np.nan
            df["force_N"]    = df["load_raw"].to_numpy() * scale
            df["stress_MPa"] = df["force_N"] / area if np.isfinite(area) else np.nan
        # else: old L2 CSV already has force_N / stress_MPa — use as-is

        df = truncate_df(df)

        # Keep pre-smoothing copies (still truncated) for the gray
        # diagnostic overlay on the plot.
        df["stress_MPa_raw"]   = df["stress_MPa"]
        df["strain_axial_raw"] = df["strain_axial"]

        area = float(df["area_mm2"].iloc[0]) if "area_mm2" in df.columns else np.nan
        df["force_N"]      = smooth_signal(df["force_N"].to_numpy())
        df["stress_MPa"]   = df["force_N"] / area if np.isfinite(area) else np.nan
        df["strain_axial"] = smooth_signal(df["strain_axial"].to_numpy())
        if "strain_transverse" in df.columns:
            df["strain_transverse"] = smooth_signal(df["strain_transverse"].to_numpy())

        p = compute_properties(df)
        props_by_cid[cid] = p

        if p:
            print(f"[{cid}]  E={p['E_GPa']:.2f} GPa  "
                  f"σ_y={p['sigma_y_MPa']:.1f} MPa  "
                  f"UTS={p['UTS_MPa']:.1f} MPa  "
                  f"ε_UTS={p['eps_at_UTS']*100:.2f}%  "
                  f"ν_chord={p['poisson_chord']:.3f}  "
                  f"toe={p['eps_toe']*100:.3f}%")
            fig_dir = FIGS_ROOT / cid
            fig_dir.mkdir(parents=True, exist_ok=True)
            plot_stress_strain(cid, p, fig_dir)
            plot_poisson(cid, p, fig_dir)

            # ---- write _L3.csv to DIC_DIR for MATLAB group figures -------
            # Per-frame signals only (toe-corrected). Scalar properties are
            # NOT repeated here — they live once per coupon in
            # level3_summary.csv and in SPECIMEN_SHEET (see write_specimen_sheet).
            n_frames = len(p["_eps"])
            pd.DataFrame({
                "step":  np.arange(n_frames),
                "eps":   p["_eps"],
                "sig":   p["_sig"],
                "eps_t": p["_eps_t"],
                "i_uts": p["_i_uts"],
            }).to_csv(DIC_DIR / f"{cid}_L3.csv", index=False, float_format="%.6g")

            # store summary row (drop internal arrays)
            rows.append({"coupon": cid,
                         **{k: v for k, v in p.items() if not k.startswith("_")}})
        else:
            print(f"[{cid}]  insufficient data")

    if rows:
        df_sum = pd.DataFrame(rows)
        df_sum["exposure"]  = df_sum["coupon"].map(lambda c: parse_id(c)[0])
        df_sum["direction"] = df_sum["coupon"].map(lambda c: parse_id(c)[1])
        # write to both FIGS_ROOT (legacy) and DIC_DIR (for MATLAB)
        for dest in (FIGS_ROOT, DIC_DIR):
            df_sum.to_csv(dest / "level3_summary.csv", index=False)

        # ---- D638 §11.7 / §12.1: mean & std per (exposure, direction) -------
        agg_cols = ["E_GPa", "sigma_y_MPa", "UTS_MPa", "eps_at_UTS", "poisson_chord"]
        group = (df_sum.groupby(["exposure", "direction"])[agg_cols]
                       .agg(["mean", "std", "count"]))
        for dest in (FIGS_ROOT, DIC_DIR):
            group.to_csv(dest / "level3_group_stats.csv")

        write_specimen_sheet(rows)

        # Group figures moved to tensile_group_plots.py
        print(f"\nCSVs written -> {FIGS_ROOT}  (run tensile_group_plots.py for group figures)")

    print(f"\nDone. {time.time()-t0:.1f} s")


if __name__ == "__main__":
    main()
