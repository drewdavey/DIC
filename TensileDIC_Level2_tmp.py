#!/usr/bin/env python3
"""
DIC_Level2_tmp.py  —  FSR Tensile Coupons (one-off batch run)
===============================================================
Production Level-2 run for THIS batch only, using peak-force-anchored raw
MTS force/displacement instead of DIC_Level2.py's load_raw*scale approach.

Why this exists: the DIC sync CSV's own "Load" channel for this batch is
only usable as a *shape*, rescaled to the true MTS peak (DIC_Level2.py's
method). Here we instead map the full raw MTS force/displacement time
series onto each DIC frame by treating the DIC-sync peak-force row and the
raw-MTS peak-force row as the same physical instant (they don't share a
clock), then interpolating MTS force/displacement at each DIC frame's time
offset from that anchor. This uses the trustworthy MTS channels for the
whole curve, not just its peak.

This script writes the same production outputs as DIC_Level2.py so
DIC_Level3.py can consume them unchanged. DIC_Level2.py itself is left
untouched — it remains the standard pipeline for future batches.

OUTPUT per coupon (identical contract to DIC_Level2.py, plus one extra
column of its own)
  <DIC_DIR>/<coupon_id>.csv   Level-1's columns, unchanged, plus:
                               kept, force_N, stress_MPa, strain_axial,
                               strain_transverse, stress_MPa_unsmoothed,
                               strain_axial_unsmoothed, disp_mm_mts (the
                               peak-anchored MTS displacement this script
                               derives — kept separate from Level-1's own
                               disp_mm, which stays untouched as the raw
                               sync-CSV reference).
  FSR-SpecimenTesting.xlsx   scalar properties written into each coupon's row
  <DIC_DIR>/level2_group_stats.csv   D638 §11.7 mean/std/count by exposure×direction
"""

from __future__ import annotations

import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.ndimage import median_filter
from scipy.signal import butter, filtfilt
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# =============================================================================
# PATHS
# =============================================================================
DIC_DIR = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\04_TestCoupons"
    r"\P01-LT150-LH4.5\DIC"
)
MTS_DIR = DIC_DIR.parent / "MTS"
SPECIMEN_SHEET = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\FSR-SpecimenTesting.xlsx"
)

# Raw DIC-sync CSVs (unreliable displacement channel, but their Time column
# and force *shape* are what we use to locate the peak-force row per coupon).
DATA_ROOTS = {
    "CL": DIC_DIR / "raw" / "2026_FSR_TensileTest_TCL",
    "SW": DIC_DIR / "raw" / "2026_FSR_TensileTest_TSW_TIS_TUV",
    "UV": DIC_DIR / "raw" / "2026_FSR_TensileTest_TSW_TIS_TUV",
    "IS": DIC_DIR / "raw" / "2026_FSR_TensileTest_TSW_TIS_TUV",
}

# =============================================================================
# SWITCHES
# =============================================================================
PRINTS     = ["P01"]
EXPOSURES  = {"CL": True, "UV": True, "SW": True, "IS": True}
DIRECTIONS = {"00": True, "45": True, "90": True}
REPLICATES = ["01", "02", "03"]

APPLY_SMOOTHING = False  # ASTM D638 does not call for filtering the stress-strain record;
                          # toggle on only if a batch's raw signal is genuinely too noisy.
FILTER_METHOD   = "butterworth"  # "median" or "butterworth" — see SMOOTHING section below

# =============================================================================
# CONSTANTS / COLUMNS — DIC-sync raw file
# =============================================================================
HEADERS = 8
IN2MM   = 25.4

DIC_FORCE_SCALED_COL = "LOAD_[kip]_|_CH07_/ai2_scaled"

# =============================================================================
# FAILURE TRUNCATION  — restricts Level-1 data to the valid test window
# before property extraction (mirrors DIC_Level2.py).
# =============================================================================
LOAD_START_FRAC = 0.02
LOAD_END_FRAC   = 0.50

# =============================================================================
# PROPERTY SETTINGS  (mirrors DIC_Level2.py)
# =============================================================================
MODULUS_STRAIN_RANGE = (0.0005, 0.003)
YIELD_OFFSET = 0.002
POISSON_RANGE = (0.0005, 0.0025)
POISSON_CHORD_AT = 0.002

SPECIMEN_SHEET_COLUMNS = {
    "E_GPa":         "E (GPa)",
    "eps_toe":       "Toe Strain",
    "sigma_y_MPa":   "Yield Stress (MPa)",
    "eps_y":         "Yield Strain",
    "UTS_MPa":       "UTS (MPa)",
    "eps_at_UTS":    "Strain at UTS",
    "poisson_chord": "Poisson's Ratio (chord)",
    "poisson_slope": "Poisson's Ratio (slope)",
}

# =============================================================================
# SMOOTHING  (mirrors DIC_Level2.py)
# =============================================================================
MEDIAN_WINDOW = 31

BUTTER_ORDER  = 3
BUTTER_CUTOFF = 0.1   # fraction of Nyquist (0-1) — must be < 1

# =============================================================================
# HELPERS
# =============================================================================
def coupon_id(p, e, d, r): return f"{p}-T{e}{d}-{r}"

def selected_coupons():
    return [coupon_id(p, e, d, r)
            for p in PRINTS
            for e, on in EXPOSURES.items() if on
            for d, on2 in DIRECTIONS.items() if on2
            for r in REPLICATES]

def parse_id(cid):
    part = cid.split("-")[1]
    return part[1:-2], part[-2:]

def coupon_dir(cid):
    exposure, _ = parse_id(cid)
    return DATA_ROOTS[exposure] / cid

def find_frames_csv(cid):
    p = DIC_DIR / f"{cid}.csv"
    return p if p.exists() else None

def load_coupon_scalars() -> pd.DataFrame:
    """coupon_scalars.csv, indexed by coupon — only area_mm2 is needed here
    (mts_peak_N isn't: force comes straight from the mapped MTS record)."""
    fp = DIC_DIR / "coupon_scalars.csv"
    if not fp.exists():
        return pd.DataFrame(columns=["mts_peak_N", "area_mm2"])
    return pd.read_csv(fp).set_index("coupon")

def find_mts_txt(cid):
    exact = [MTS_DIR / f"{cid}.txt", MTS_DIR / f"{cid}-TEST.txt"]
    for fp in exact:
        if fp.exists():
            return fp
    hits = sorted(MTS_DIR.glob(f"{cid}*.txt"))
    return hits[0] if hits else None

def find_dic_sync_csv(cid):
    cdir = coupon_dir(cid)
    direct = cdir / f"{cid}.csv"
    if direct.exists():
        return direct
    if not cdir.is_dir():
        return None
    hits = sorted(cdir.rglob(f"{cid}.csv"))
    return hits[0] if hits else None

def pick_col(df, hint):
    for col in df.columns:
        if hint.lower() in str(col).lower():
            return col
    return None

def numeric(values):
    return pd.to_numeric(values, errors="coerce").to_numpy(dtype=float)

def relative_time(values, fallback_len):
    arr = numeric(values)
    finite = np.flatnonzero(np.isfinite(arr))
    if finite.size:
        return arr - arr[finite[0]]
    return np.arange(fallback_len, dtype=float)

def sample_rate_hz(time_s):
    dt = np.diff(time_s[np.isfinite(time_s)])
    dt = dt[np.isfinite(dt) & (dt > 0)]
    if not dt.size:
        return np.nan
    return float(1.0 / np.median(dt))

def load_mts_txt(fp):
    return (
        pd.read_csv(
            fp,
            sep="\t",
            skiprows=HEADERS,
            header=None,
            names=["disp_mm", "force_N", "output_V", "time_s"],
            encoding="utf-8-sig",
            on_bad_lines="skip",
        )
        .apply(pd.to_numeric, errors="coerce")
        .dropna(how="all")
    )

def smooth_signal(x):
    x = np.asarray(x, dtype=float)
    if not APPLY_SMOOTHING:
        return x.copy()
    if FILTER_METHOD == "butterworth":
        return _smooth_butterworth(x)
    return _smooth_median(x)

def _smooth_median(x):
    win = MEDIAN_WINDOW
    if win % 2 == 0:
        win -= 1
    if win < 1 or len(x) < win:
        return x.copy()
    return median_filter(x, size=win, mode="constant", cval=np.nan)

def _smooth_butterworth(x):
    n = len(x)
    padlen = 3 * BUTTER_ORDER
    nan_mask = ~np.isfinite(x)
    if nan_mask.all() or n <= padlen:
        return x.copy()
    xi = x.copy()
    if nan_mask.any():
        idx = np.arange(n)
        xi[nan_mask] = np.interp(idx[nan_mask], idx[~nan_mask], x[~nan_mask])
    b, a = butter(BUTTER_ORDER, BUTTER_CUTOFF, btype="low", analog=False)
    out = filtfilt(b, a, xi)
    out = np.clip(out, np.nanmin(xi), np.nanmax(xi))
    out[nan_mask] = np.nan
    return out

def write_specimen_sheet(rows: list[dict]) -> None:
    try:
        wb = openpyxl.load_workbook(SPECIMEN_SHEET)
    except FileNotFoundError:
        print(f"[!] {SPECIMEN_SHEET} not found — skipping specimen sheet update")
        return
    ws = wb.active

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    id_col = header.get("Specimen ID")
    if id_col is None:
        print("[!] 'Specimen ID' column not found in specimen sheet — skipping update")
        return

    next_col = ws.max_column + 1
    for label in SPECIMEN_SHEET_COLUMNS.values():
        if label not in header:
            ws.cell(row=1, column=next_col, value=label)
            header[label] = next_col
            next_col += 1

    row_by_id = {ws.cell(row=r, column=id_col).value: r
                 for r in range(2, ws.max_row + 1)}

    for row in rows:
        r = row_by_id.get(row["coupon"])
        if r is None:
            continue
        for key, label in SPECIMEN_SHEET_COLUMNS.items():
            v = row.get(key)
            v = None if (v is None or not np.isfinite(v)) else v
            ws.cell(row=r, column=header[label], value=v)

    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(SPECIMEN_SHEET)
    except PermissionError:
        print(f"[!] {SPECIMEN_SHEET} is open elsewhere — could not save properties to it")

def truncation_mask(force_N: np.ndarray) -> np.ndarray:
    """Boolean mask, same length as force_N, selecting the valid test
    window: pre-load slack and post-fracture rebound are False."""
    n = len(force_N)
    peak = float(np.nanmax(np.abs(force_N)))
    if peak <= 0:
        return np.ones(n, dtype=bool)
    i_uts = int(np.nanargmax(np.abs(force_N)))
    starts = np.where(np.abs(force_N) > LOAD_START_FRAC * peak)[0]
    i0 = int(starts[0]) if len(starts) else 0
    post = np.where(np.abs(force_N[i_uts:]) < LOAD_END_FRAC * peak)[0]
    i1 = int(i_uts + post[0]) - 1 if len(post) else n - 1
    mask = np.zeros(n, dtype=bool)
    mask[i0:i1 + 1] = True
    return mask


# =============================================================================
# PEAK-FORCE-ANCHORED MTS MAPPING
# =============================================================================
def map_mts_force_disp(cid, n_l1):
    """Map raw MTS force/displacement onto DIC-frame times.

    The DIC sync CSV and the raw MTS file don't share a clock, so the
    DIC-sync peak-force row and the raw-MTS peak-force row are treated as
    the same physical instant. Each DIC frame's time offset from that
    anchor is used to interpolate MTS force/displacement at the
    corresponding MTS-clock time. Returns (force_N, disp_mm), each length
    n_l1 (NaN-padded past the sync CSV's length if it's shorter), or None.
    """
    sync_fp = find_dic_sync_csv(cid)
    mts_fp = find_mts_txt(cid)
    if sync_fp is None or mts_fp is None:
        print(f"[{cid}] skip: missing sync={bool(sync_fp)} MTS={bool(mts_fp)}")
        return None

    sync = pd.read_csv(sync_fp)
    time_col = pick_col(sync, "time")
    if DIC_FORCE_SCALED_COL not in sync.columns or time_col is None:
        print(f"[{cid}] skip: sync CSV missing force or time column")
        return None

    n = min(n_l1, len(sync))
    sync = sync.iloc[:n].reset_index(drop=True)
    sync_time = relative_time(sync[time_col], len(sync))
    sync_fs = sample_rate_hz(sync_time)
    sync_force_kip = numeric(sync[DIC_FORCE_SCALED_COL])
    sync_peak_i = int(np.nanargmax(np.abs(sync_force_kip)))

    mts = load_mts_txt(mts_fp)
    mts_time = relative_time(mts["time_s"], len(mts))
    mts_fs = sample_rate_hz(mts_time)
    mts_force_N = numeric(mts["force_N"])
    mts_disp_mm = numeric(mts["disp_mm"])
    mts_peak_i = int(np.nanargmax(np.abs(mts_force_N)))

    dic_dt_from_peak = sync_time - sync_time[sync_peak_i]
    mts_mapped_time = mts_time[mts_peak_i] + dic_dt_from_peak

    mapped_force_N = np.interp(mts_mapped_time, mts_time, mts_force_N, left=np.nan, right=np.nan)
    mapped_disp_mm = np.interp(mts_mapped_time, mts_time, mts_disp_mm, left=np.nan, right=np.nan)
    finite_disp = np.flatnonzero(np.isfinite(mapped_disp_mm))
    if finite_disp.size:
        mapped_disp_mm = mapped_disp_mm - mapped_disp_mm[finite_disp[0]]

    print(f"[{cid}] sync_fs={sync_fs:.3f} Hz  mts_fs={mts_fs:.2f} Hz  "
          f"peak_i sync/MTS={sync_peak_i}/{mts_peak_i}  "
          f"MTS_peak={abs(mts_force_N[mts_peak_i]):.1f} N")

    if n < n_l1:
        pad = np.full(n_l1 - n, np.nan)
        mapped_force_N = np.concatenate([mapped_force_N, pad])
        mapped_disp_mm = np.concatenate([mapped_disp_mm, pad])
    return mapped_force_N, mapped_disp_mm


# =============================================================================
# COMPUTE PROPERTIES  (mirrors DIC_Level2.py)
# =============================================================================
def compute_properties(eps_axial, sig, eps_transverse,
                       eps_axial_unsmoothed=None, sig_unsmoothed=None):
    eps_axial = np.asarray(eps_axial, dtype=float)
    sig = np.asarray(sig, dtype=float)
    eps_transverse = np.asarray(eps_transverse, dtype=float)

    valid = np.isfinite(eps_axial) & np.isfinite(sig)
    if valid.sum() < 10:
        return None

    eps_raw   = eps_axial[valid]
    sig       = sig[valid]
    eps_t_raw = eps_transverse[valid]
    eps_unsmoothed = (np.asarray(eps_axial_unsmoothed, dtype=float)[valid]
                       if eps_axial_unsmoothed is not None else None)
    sig_unsmoothed = (np.asarray(sig_unsmoothed, dtype=float)[valid]
                       if sig_unsmoothed is not None else None)

    lo, hi = MODULUS_STRAIN_RANGE
    mfit = (eps_raw >= lo) & (eps_raw <= hi) & np.isfinite(eps_raw) & np.isfinite(sig)
    if mfit.sum() < 3:
        return None
    slope, intercept = np.polyfit(eps_raw[mfit], sig[mfit], 1)
    E_MPa = float(slope)

    eps_offset = -intercept / E_MPa if E_MPa != 0 else 0.0
    eps   = eps_raw   - eps_offset
    eps_unsmoothed_corr = (eps_unsmoothed - eps_offset) if eps_unsmoothed is not None else None
    if np.any(np.isfinite(eps_t_raw)):
        i0 = int(np.nanargmin(np.abs(eps)))
        eps_t = eps_t_raw - eps_t_raw[i0]
    else:
        eps_t = eps_t_raw.copy()

    i_uts   = int(np.nanargmax(sig))
    uts     = float(sig[i_uts])
    eps_ult = float(eps[i_uts])

    sigma_y, eps_y = np.nan, np.nan
    diff  = sig - E_MPa * (eps - YIELD_OFFSET)
    valid_y = np.where(eps > YIELD_OFFSET)[0]
    if len(valid_y) > 1:
        d = diff[valid_y]
        crossings = np.where(np.diff(np.sign(d)) < 0)[0]
        if len(crossings):
            k = valid_y[crossings[0]]
            denom = diff[k] - diff[k+1]
            f = diff[k] / denom if denom != 0 else 0.0
            eps_y   = float(eps[k] + f * (eps[k+1] - eps[k]))
            sigma_y = float(sig[k] + f * (sig[k+1] - sig[k]))

    nu_chord = nu_slope = np.nan
    pm = ((eps >= POISSON_RANGE[0]) & (eps <= POISSON_RANGE[1]) &
          np.isfinite(eps) & np.isfinite(eps_t))
    if pm.sum() >= 3:
        nu_slope = float(-np.polyfit(eps[pm], eps_t[pm], 1)[0])
        order = np.argsort(eps[pm])
        ea, et = eps[pm][order], eps_t[pm][order]
        if ea[0] <= POISSON_CHORD_AT <= ea[-1]:
            nu_chord = float(-np.interp(POISSON_CHORD_AT, ea, et) / POISSON_CHORD_AT)

    return {
        "E_GPa":         E_MPa / 1000.0,
        "eps_toe":       eps_offset,
        "sigma_y_MPa":   sigma_y,
        "eps_y":         eps_y,
        "UTS_MPa":       uts,
        "eps_at_UTS":    eps_ult,
        "poisson_chord": nu_chord,
        "poisson_slope": nu_slope,
        "_eps":          eps,
        "_sig":          sig,
        "_eps_t":        eps_t,
        "_eps_raw":      eps_unsmoothed_corr,
        "_sig_raw":      sig_unsmoothed,
        "_valid":        valid,
    }


# =============================================================================
# MAIN
# =============================================================================
def main():
    t0 = time.time()
    print("=" * 70)
    print("DIC_Level2_tmp — batch Level-2 run using peak-anchored MTS force mapping")
    print("=" * 70)
    rows = []
    scalars = load_coupon_scalars()

    for cid in selected_coupons():
        frames_fp = find_frames_csv(cid)
        if frames_fp is None:
            print(f"[{cid}] no per-coupon CSV — run Level 1 first")
            continue
        df = pd.read_csv(frames_fp)
        n = len(df)

        mapped = map_mts_force_disp(cid, n)
        if mapped is None:
            continue
        force_N_all, disp_mm_mts = mapped

        area = (float(scalars.loc[cid, "area_mm2"])
                if cid in scalars.index and pd.notna(scalars.loc[cid, "area_mm2"])
                else np.nan)

        # ---- failure truncation ---------------------------------------------
        kept = truncation_mask(force_N_all)
        kept_idx = np.flatnonzero(kept)

        force_kept  = force_N_all[kept]
        stress_kept = force_kept / area if np.isfinite(area) else np.full(kept.sum(), np.nan)
        eps_a_kept  = df["strain_axial_raw"].to_numpy()[kept]
        eps_t_kept  = df["strain_transverse_raw"].to_numpy()[kept]

        force_smoothed  = smooth_signal(force_kept)
        stress_smoothed = force_smoothed / area if np.isfinite(area) else np.full(kept.sum(), np.nan)
        eps_a_smoothed  = smooth_signal(eps_a_kept)
        eps_t_smoothed  = smooth_signal(eps_t_kept)

        p = compute_properties(eps_a_smoothed, stress_smoothed, eps_t_smoothed,
                               eps_axial_unsmoothed=eps_a_kept, sig_unsmoothed=stress_kept)
        if not p:
            print(f"[{cid}]  insufficient data")
            continue

        print(f"[{cid}]  E={p['E_GPa']:.2f} GPa  "
              f"σ_y={p['sigma_y_MPa']:.1f} MPa  "
              f"UTS={p['UTS_MPa']:.1f} MPa  "
              f"ε_UTS={p['eps_at_UTS']*100:.2f}%  "
              f"ν_chord={p['poisson_chord']:.3f}  "
              f"toe={p['eps_toe']*100:.3f}%")

        # ---- scatter L2 columns back onto the full per-frame file -----------
        final_idx = kept_idx[p["_valid"]]

        def scatter(values):
            col = np.full(n, np.nan)
            col[final_idx] = values
            return col

        df["kept"]                     = kept
        df["force_N"]                  = scatter(force_smoothed[p["_valid"]])
        df["stress_MPa"]               = scatter(p["_sig"])
        df["strain_axial"]             = scatter(p["_eps"])
        df["strain_transverse"]        = scatter(p["_eps_t"])
        df["stress_MPa_unsmoothed"]    = scatter(p["_sig_raw"])
        df["strain_axial_unsmoothed"]  = scatter(p["_eps_raw"])
        df["disp_mm_mts"]              = disp_mm_mts
        df.to_csv(frames_fp, index=False, float_format="%.6g")

        rows.append({"coupon": cid, **{k: v for k, v in p.items() if not k.startswith("_")}})

    if rows:
        write_specimen_sheet(rows)

        df_sum = pd.DataFrame(rows)
        df_sum["exposure"]  = df_sum["coupon"].map(lambda c: parse_id(c)[0])
        df_sum["direction"] = df_sum["coupon"].map(lambda c: parse_id(c)[1])
        agg_cols = ["E_GPa", "sigma_y_MPa", "UTS_MPa", "eps_at_UTS", "poisson_chord"]
        group = (df_sum.groupby(["exposure", "direction"])[agg_cols]
                       .agg(["mean", "std", "count"]))
        group.to_csv(DIC_DIR / "level2_group_stats.csv")

        print(f"\n{len(rows)} coupon(s) → DIC/*.csv, {SPECIMEN_SHEET.name}, "
              f"DIC/level2_group_stats.csv")

    print(f"\nDone. {time.time()-t0:.1f} s")


if __name__ == "__main__":
    main()
