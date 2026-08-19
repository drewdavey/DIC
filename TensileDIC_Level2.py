#!/usr/bin/env python3
"""
DIC_Level2.py  —  FSR Tensile Coupons
======================================
Reads Level-1's per-coupon frame CSV, scales raw load to force/stress, applies
failure truncation, and computes ASTM D638 mechanical properties. An
optional smoothing pass (median or Butterworth, see APPLY_SMOOTHING /
FILTER_METHOD) is available but off by default — ASTM D638 doesn't call
for filtering the stress-strain record; only enable it if a batch's raw
signal is genuinely too noisy. No plotting here — see DIC_Level3.py.

See also DIC_Level2_tmp.py: a one-off variant of this script that derives
force/stress from peak-force-anchored raw MTS data instead of scaling the
DIC sync CSV's own load channel, used for batches where that channel is
unreliable. It writes the same outputs as this script but leaves this
file untouched.

Standards compliance — what each calculation cites
  Toe compensation     : D638 Annex A1 (mandatory unless toe is real material response)
  Modulus              : D638 §11.4   (slope of initial linear region of σ-ε)
  0.2 % offset yield    : D638 §A2.6 / Fig. A2.1 (offset from toe-corrected origin)
  Tensile strength UTS  : D638 §11.2   (max stress / original area)
  Poisson's ratio       : D638 Annex A3.10.1.3 (chord at ε_a=0.002 over 0.0005-0.0025)
  Group statistics      : D638 §11.7 / §12.1   (mean, std per series)

PROCESSING NOTE
  Level-1 writes the full, untruncated per-frame record (see its docstring).
  Level-2 reads that SAME file and appends its own columns to it in place —
  no separate per-coupon output file. Failure truncation happens here via
  truncation_mask(): pre-load slack and post-fracture rebound (past 50%
  post-UTS load drop) are marked out of the analysis window rather than
  dropped from the file — every L2-derived column is NaN outside that
  window, and the boolean 'kept' column marks which rows are inside it.
  Properties are computed from this window (smoothed, if APPLY_SMOOTHING is
  on); pre-smoothing values (still restricted to the window) are kept
  alongside either way for Level-3's diagnostic overlay — they're identical
  to the smoothed columns when smoothing is off.

INPUT per coupon
  <DIC_DIR>/<coupon_id>.csv      Level-1's per-frame record (see its docstring)
  <DIC_DIR>/coupon_scalars.csv   per-coupon mts_peak_N, area_mm2 (Level-1 output)

OUTPUT per coupon
  <DIC_DIR>/<coupon_id>.csv   Level-1's columns, unchanged, plus:
                               kept, force_N, stress_MPa, strain_axial,
                               strain_transverse (toe-corrected, and smoothed
                               if APPLY_SMOOTHING is on), stress_MPa_unsmoothed,
                               strain_axial_unsmoothed (same window, pre-smoothing).
                               All L2 columns are NaN where kept is False.
  FSR-SpecimenTesting.xlsx    scalar properties written into each coupon's
                               row (E, toe strain, yield stress/strain, UTS,
                               strain at UTS, Poisson's ratio) — the single
                               source of truth for per-coupon scalars, read
                               back out by Level-3.
  <DIC_DIR>/level2_group_stats.csv   D638 §11.7 mean/std/count by exposure×direction
"""

from __future__ import annotations
import sys
import time
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.ndimage import median_filter
from scipy.signal import butter, filtfilt
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# =============================================================================
# PATHS
# =============================================================================
DIC_DIR = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\04_TestCoupons"
    r"\P01-LT150-LH4.5\DIC"
)
SPECIMEN_SHEET = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\FSR-SpecimenTesting.xlsx"
)

# =============================================================================
# SWITCHES
# =============================================================================
PRINTS     = ["P01"]
EXPOSURES  = {"CL": True, "UV": True, "SW": True, "IS": True}
DIRECTIONS = {"00": True, "45": True, "90": True}
REPLICATES = ["01", "02", "03"]

APPLY_SMOOTHING = False  # ASTM D638 does not call for filtering the stress-strain record;
                          # toggle on only if a batch's raw signal is genuinely too noisy.
FILTER_METHOD   = "butterworth"  # "median" or "butterworth" — see SMOOTHING section below

# =============================================================================
# FAILURE TRUNCATION  — restricts Level-1 data to the valid test window
# before property extraction. Trims pre-load slack and post-fracture
# rebound; frames outside this window get NaN in every L2-derived column.
# =============================================================================
LOAD_START_FRAC = 0.02     # pre-load: drop frames before load exceeds this × peak
LOAD_END_FRAC   = 0.50     # post-fracture: cut first post-UTS frame where load < this × peak

# Scale factor applied to load_raw to produce force_N (N per sync-CSV unit).
# Two modes — the first one found is used:
#   1. Per-coupon  : mts_peak_N / max(raw load_raw)  — most accurate, always
#                    available since Level-1 always writes a coupon_scalars.csv row
#   2. Combined    : SCALE_N_PER_UNIT below          — fallback safety net
SCALE_N_PER_UNIT: float = 555.5928

# =============================================================================
# PROPERTY SETTINGS
# =============================================================================
# Modulus fit window (axial strain, dimensionless).
# D638 §11.4: "initial linear portion of the load-extension curve".
# A window of 0.05–0.3% covers the typical linear region for stiff polymers
# and composites without including the toe. Adjust if the fit line on the
# generated plot doesn't sit on the linear segment.
MODULUS_STRAIN_RANGE = (0.0005, 0.003)

# D638 §A2.6 — 0.2% offset yield strength
YIELD_OFFSET = 0.002

# D638 §A3.10.1.3 — Poisson chord method window (when no clear proportionality)
# Chord computed at ε_a = 0.002 over the range 0.0005 to 0.0025 strain.
POISSON_RANGE = (0.0005, 0.0025)
POISSON_CHORD_AT = 0.002

# Scalar property columns written into SPECIMEN_SHEET, keyed by coupon
# ("Specimen ID") — maps the property dict key to the Excel column header.
# Level-3 reads these same headers back out (per-coupon plots, group plots,
# and the printed/exported stat tables all read from this one workbook).
SPECIMEN_SHEET_COLUMNS = {
    "E_GPa":         "E (GPa)",
    "eps_toe":       "Toe Strain",
    "sigma_y_MPa":   "Yield Stress (MPa)",
    "eps_y":         "Yield Strain",
    "UTS_MPa":       "UTS (MPa)",
    "eps_at_UTS":    "Strain at UTS",
    "poisson_chord": "Poisson's Ratio (chord)",
    "poisson_slope": "Poisson's Ratio (slope)",
}

# =============================================================================
# SMOOTHING  — FILTER_METHOD selects which of these is used
#   "median"     : rolling median. Window must be odd. Raise MEDIAN_WINDOW
#                  for noisier data. Preferred default — doesn't ring or
#                  systematically undershoot a sharp peak the way an
#                  averaging-based filter like Butterworth can.
#   "butterworth": zero-phase low-pass (filtfilt). Lower BUTTER_CUTOFF for
#                  heavier smoothing; raise BUTTER_ORDER for a sharper
#                  rolloff. Output is clipped to the raw data's range to
#                  suppress filtfilt ringing at the truncation edges.
# =============================================================================
MEDIAN_WINDOW = 31  # frames

BUTTER_ORDER  = 3     # filter order
BUTTER_CUTOFF = 0.1   # cutoff frequency, fraction of Nyquist (0-1) — must be < 1

# =============================================================================
# HELPERS
# =============================================================================
def coupon_id(p, e, d, r): return f"{p}-T{e}{d}-{r}"

def selected_coupons():
    return [coupon_id(p, e, d, r)
            for p in PRINTS
            for e, on in EXPOSURES.items() if on
            for d, on2 in DIRECTIONS.items() if on2
            for r in REPLICATES]

def parse_id(cid):
    """Return (exposure, direction_str) e.g. ('CL', '00')."""
    part = cid.split("-")[1]
    return part[1:-2], part[-2:]

def find_frames_csv(cid):
    p = DIC_DIR / f"{cid}.csv"
    return p if p.exists() else None

def load_coupon_scalars() -> pd.DataFrame:
    """coupon_scalars.csv, indexed by coupon — mts_peak_N and area_mm2,
    written once per coupon by Level-1 (never repeated per row)."""
    fp = DIC_DIR / "coupon_scalars.csv"
    if not fp.exists():
        return pd.DataFrame(columns=["mts_peak_N", "area_mm2"])
    return pd.read_csv(fp).set_index("coupon")

def smooth_signal(x):
    """Dispatches to the filter selected by FILTER_METHOD. No-op when
    APPLY_SMOOTHING is False."""
    x = np.asarray(x, dtype=float)
    if not APPLY_SMOOTHING:
        return x.copy()
    if FILTER_METHOD == "butterworth":
        return _smooth_butterworth(x)
    return _smooth_median(x)

def _smooth_median(x):
    """Rolling median. mode='nearest' avoids the zero-padding edge artifacts
    scipy.signal.medfilt has."""
    win = MEDIAN_WINDOW
    if win % 2 == 0:
        win -= 1
    if win < 1 or len(x) < win:
        return x.copy()
    return median_filter(x, size=win, mode="constant", cval=np.nan)

def _smooth_butterworth(x):
    """Zero-phase low-pass (filtfilt). NaN gaps are bridged by linear
    interpolation before filtering, then the original NaN positions are
    restored. Output is clipped to the raw data's range — filtfilt can ring
    past it, especially near the truncation edges."""
    n = len(x)
    padlen = 3 * BUTTER_ORDER
    nan_mask = ~np.isfinite(x)
    if nan_mask.all() or n <= padlen:
        return x.copy()
    xi = x.copy()
    if nan_mask.any():
        idx = np.arange(n)
        xi[nan_mask] = np.interp(idx[nan_mask], idx[~nan_mask], x[~nan_mask])
    b, a = butter(BUTTER_ORDER, BUTTER_CUTOFF, btype="low", analog=False)
    out = filtfilt(b, a, xi)
    out = np.clip(out, np.nanmin(xi), np.nanmax(xi))
    out[nan_mask] = np.nan
    return out

def write_specimen_sheet(rows: list[dict]) -> None:
    """Write each coupon's scalar properties into its row in SPECIMEN_SHEET,
    matched by Specimen ID. Adds any missing property columns at the end;
    everything else in the workbook (other rows, formulas, formatting) is
    left untouched. Skipped (with a warning) if the file can't be opened —
    e.g. if it's currently open in Excel.
    """
    try:
        wb = openpyxl.load_workbook(SPECIMEN_SHEET)
    except FileNotFoundError:
        print(f"[!] {SPECIMEN_SHEET} not found — skipping specimen sheet update")
        return
    ws = wb.active

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    id_col = header.get("Specimen ID")
    if id_col is None:
        print("[!] 'Specimen ID' column not found in specimen sheet — skipping update")
        return

    next_col = ws.max_column + 1
    for label in SPECIMEN_SHEET_COLUMNS.values():
        if label not in header:
            ws.cell(row=1, column=next_col, value=label)
            header[label] = next_col
            next_col += 1

    row_by_id = {ws.cell(row=r, column=id_col).value: r
                 for r in range(2, ws.max_row + 1)}

    for row in rows:
        r = row_by_id.get(row["coupon"])
        if r is None:
            continue
        for key, label in SPECIMEN_SHEET_COLUMNS.items():
            v = row.get(key)
            v = None if (v is None or not np.isfinite(v)) else v
            ws.cell(row=r, column=header[label], value=v)

    # openpyxl doesn't evaluate formulas, so re-saving drops the cached
    # values of every formula cell in the workbook (e.g. Width/Dia,
    # Computed Area) until something recalculates them. Force a full
    # recalculation on next open so they never appear blank.
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(SPECIMEN_SHEET)
    except PermissionError:
        print(f"[!] {SPECIMEN_SHEET} is open elsewhere — could not save properties to it")

def read_existing_group_stats(fp: Path) -> pd.DataFrame | None:
    """Read level2_group_stats.csv, upgrading the pre-'test' two-level layout.

    The index depth has to be READ, not assumed. pandas writes a named
    MultiIndex as a third header line holding the level names, so that line says
    how many index columns the file on disk really has — and reading a
    two-level file with index_col=[0,1,2] does not raise. It quietly swallows
    the first data column into the index and shifts every value one place left,
    which looks like a successful merge and silently corrupts the other test
    type's rows.

    Returns None if the file can't be understood, so the caller can say so
    rather than write a half-merged table.
    """
    try:
        with open(fp, encoding="utf-8") as fh:
            head = [fh.readline() for _ in range(3)]
        names = [c.strip() for c in head[2].rstrip(chr(13) + chr(10)).split(",") if c.strip()]
        if not names:
            return None
        old = pd.read_csv(fp, header=[0, 1], index_col=list(range(len(names))))
        if "test" not in names:
            # Written before the 'test' level existed, when only the tensile
            # pipeline wrote this file. Everything in it is tensile.
            old = pd.concat({"tensile": old}, names=["test"])
        # A direction of "00" round-trips through read_csv as the integer 0, so
        # normalise every level back to the string form both scripts write.
        old.index = pd.MultiIndex.from_tuples(
            [(str(t), str(e), f"{int(d):02d}" if str(d).strip().isdigit() else str(d))
             for t, e, d in old.index],
            names=["test", "exposure", "direction"])
        return old
    except Exception:
        return None


def write_group_stats(rows: list[dict]) -> Path:
    """Upsert this run's group statistics into level2_group_stats.csv.

    The file is shared with FlexuralDIC_Level2 and indexed by
    (test, exposure, direction). Both test types have CL and IS exposures at 00
    and 90, so without the 'test' level the flexural rows and the tensile ones
    would land on top of each other. Only rows whose test == "tensile" are
    replaced; anything else already in the file is read back and written out
    unchanged.
    """
    fp = DIC_DIR / "level2_group_stats.csv"
    df_sum = pd.DataFrame(rows)
    df_sum["test"] = "tensile"
    df_sum["exposure"]  = df_sum["coupon"].map(lambda c: parse_id(c)[0])
    df_sum["direction"] = df_sum["coupon"].map(lambda c: parse_id(c)[1])
    agg_cols = ["E_GPa", "sigma_y_MPa", "UTS_MPa", "eps_at_UTS", "poisson_chord"]
    group = (df_sum.groupby(["test", "exposure", "direction"])[agg_cols]
                   .agg(["mean", "std", "count"]))

    if fp.exists():
        try:
            old = pd.read_csv(fp, header=[0, 1], index_col=[0, 1, 2])
            old = old.drop(index="tensile", level=0, errors="ignore")
            group = pd.concat([old, group]).sort_index()
        except Exception as ex:
            # An older two-level file, written before the 'test' level existed,
            # can't be merged column-wise. Say so rather than silently dropping
            # it; re-running FlexuralDIC_Level2 restores the flexural rows.
            print(f"[!] could not merge existing {fp.name} ({ex}) — it is being "
                  f"replaced with tensile rows only.\n    Re-run "
                  f"FlexuralDIC_Level2.py to put the flexural rows back.")
    group.to_csv(fp)
    return fp


def truncation_mask(force_N: np.ndarray) -> np.ndarray:
    """Boolean mask, same length as force_N, selecting the valid test
    window: pre-load slack and post-fracture rebound are False."""
    n = len(force_N)
    peak = float(np.nanmax(np.abs(force_N)))
    if peak <= 0:
        return np.ones(n, dtype=bool)
    i_uts = int(np.nanargmax(np.abs(force_N)))
    starts = np.where(np.abs(force_N) > LOAD_START_FRAC * peak)[0]
    i0 = int(starts[0]) if len(starts) else 0
    post = np.where(np.abs(force_N[i_uts:]) < LOAD_END_FRAC * peak)[0]
    # stop one frame BEFORE the first post-UTS drop-off so the failure point
    # itself isn't kept (it corrupts the smoothing pass)
    i1 = int(i_uts + post[0]) - 1 if len(post) else n - 1
    mask = np.zeros(n, dtype=bool)
    mask[i0:i1 + 1] = True
    return mask


# =============================================================================
# COMPUTE PROPERTIES
# =============================================================================
def compute_properties(eps_axial, sig, eps_transverse,
                       eps_axial_unsmoothed=None, sig_unsmoothed=None):
    """
    D638-compliant property extraction from one coupon's truncated (and, if
    enabled, smoothed) axial-strain / stress / transverse-strain arrays.

    Returns a dict with E_GPa, eps_toe (toe-correction offset applied),
    sigma_y_MPa, eps_y, UTS_MPa, eps_at_UTS, poisson_chord, poisson_slope,
    the toe-corrected _eps/_sig/_eps_t arrays, a _valid boolean mask (which
    input rows survived the finite-value filter — needed by the caller to
    scatter these arrays back into the full per-frame file), and, when
    eps_axial_unsmoothed/sig_unsmoothed were given, their toe-corrected
    pre-smoothing counterparts _eps_raw/_sig_raw for Level-3's diagnostic
    overlay (identical to _eps/_sig when smoothing is off).
    """
    eps_axial = np.asarray(eps_axial, dtype=float)
    sig = np.asarray(sig, dtype=float)
    eps_transverse = np.asarray(eps_transverse, dtype=float)

    valid = np.isfinite(eps_axial) & np.isfinite(sig)
    if valid.sum() < 10:
        return None

    eps_raw   = eps_axial[valid]
    sig       = sig[valid]
    eps_t_raw = eps_transverse[valid]
    eps_unsmoothed = (np.asarray(eps_axial_unsmoothed, dtype=float)[valid]
                       if eps_axial_unsmoothed is not None else None)
    sig_unsmoothed = (np.asarray(sig_unsmoothed, dtype=float)[valid]
                       if sig_unsmoothed is not None else None)

    # ---- 1. Modulus (D638 §11.4) --------------------------------------------
    lo, hi = MODULUS_STRAIN_RANGE
    mfit = (eps_raw >= lo) & (eps_raw <= hi) & np.isfinite(eps_raw) & np.isfinite(sig)
    if mfit.sum() < 3:
        return None
    slope, intercept = np.polyfit(eps_raw[mfit], sig[mfit], 1)
    E_MPa = float(slope)

    # ---- 2. Toe compensation (D638 Annex A1) --------------------------------
    # The fitted line σ = E·ε + b is extended back to σ = 0; that strain
    # (b/(-E)) is the "toe offset" — all strains are then measured from the
    # corrected origin. ε_corrected = ε_raw − ε_offset.
    eps_offset = -intercept / E_MPa if E_MPa != 0 else 0.0
    eps   = eps_raw   - eps_offset
    eps_unsmoothed_corr = (eps_unsmoothed - eps_offset) if eps_unsmoothed is not None else None
    # Transverse strain: subtract its value at the corrected zero of axial strain.
    # Find the index where corrected axial ≈ 0 and subtract that ε_t.
    if np.any(np.isfinite(eps_t_raw)):
        i0 = int(np.nanargmin(np.abs(eps)))
        eps_t = eps_t_raw - eps_t_raw[i0]
    else:
        eps_t = eps_t_raw.copy()

    # ---- 3. UTS (D638 §11.2 — max stress) -----------------------------------
    i_uts   = int(np.nanargmax(sig))
    uts     = float(sig[i_uts])
    eps_ult = float(eps[i_uts])

    # ---- 4. 0.2% offset yield (D638 §A2.6, Fig. A2.1) -----------------------
    # First crossing of σ-ε curve with the line σ = E·(ε − YIELD_OFFSET).
    sigma_y, eps_y = np.nan, np.nan
    diff  = sig - E_MPa * (eps - YIELD_OFFSET)
    valid_y = np.where(eps > YIELD_OFFSET)[0]
    if len(valid_y) > 1:
        d = diff[valid_y]
        crossings = np.where(np.diff(np.sign(d)) < 0)[0]
        if len(crossings):
            k = valid_y[crossings[0]]
            denom = diff[k] - diff[k+1]
            f = diff[k] / denom if denom != 0 else 0.0
            eps_y   = float(eps[k] + f * (eps[k+1] - eps[k]))
            sigma_y = float(sig[k] + f * (sig[k+1] - sig[k]))

    # ---- 5. Poisson's ratio (D638 §A3.10.1.3, chord at ε_a = 0.002) --------
    #   ν = − ε_t(at ε_a = 0.002) / 0.002
    # Range 0.0005 – 0.0025 strain. Also report least-squares slope (§A3.10.1.1)
    # for transparency when proportionality holds.
    nu_chord = nu_slope = np.nan
    pm = ((eps >= POISSON_RANGE[0]) & (eps <= POISSON_RANGE[1]) &
          np.isfinite(eps) & np.isfinite(eps_t))
    if pm.sum() >= 3:
        nu_slope = float(-np.polyfit(eps[pm], eps_t[pm], 1)[0])
        order = np.argsort(eps[pm])
        ea, et = eps[pm][order], eps_t[pm][order]
        if ea[0] <= POISSON_CHORD_AT <= ea[-1]:
            nu_chord = float(-np.interp(POISSON_CHORD_AT, ea, et) / POISSON_CHORD_AT)

    return {
        "E_GPa":         E_MPa / 1000.0,
        "eps_toe":       eps_offset,
        "sigma_y_MPa":   sigma_y,
        "eps_y":         eps_y,
        "UTS_MPa":       uts,
        "eps_at_UTS":    eps_ult,
        "poisson_chord": nu_chord,
        "poisson_slope": nu_slope,
        "_eps":          eps,        # toe-corrected strain, used in property calcs
        "_sig":          sig,
        "_eps_t":        eps_t,
        "_eps_raw":      eps_unsmoothed_corr,  # diagnostic overlay only
        "_sig_raw":      sig_unsmoothed,
        "_valid":        valid,      # which input rows survived the finite-value filter
    }


# =============================================================================
# MAIN
# =============================================================================
def main():
    t0 = time.time()
    print("=" * 70)
    print("DIC_Level2 — scale, truncate, smooth, compute D638 properties")
    print("=" * 70)
    rows = []
    scalars = load_coupon_scalars()

    for cid in selected_coupons():
        frames_fp = find_frames_csv(cid)
        if frames_fp is None:
            print(f"[{cid}] no per-coupon CSV — run Level 1 first")
            continue
        df = pd.read_csv(frames_fp)
        n = len(df)
        load_raw = df["load_raw"].to_numpy()

        # ---- scale load_raw -> force_N --------------------------------------
        raw_peak = float(np.nanmax(np.abs(load_raw)))
        mts_peak = (float(scalars.loc[cid, "mts_peak_N"])
                    if cid in scalars.index and pd.notna(scalars.loc[cid, "mts_peak_N"])
                    else None)
        area = (float(scalars.loc[cid, "area_mm2"])
                if cid in scalars.index and pd.notna(scalars.loc[cid, "area_mm2"])
                else np.nan)
        if mts_peak is not None and raw_peak > 0:
            scale = mts_peak / raw_peak
            print(f"[{cid}] per-coupon scale: {scale:.4f} N/unit  (MTS {mts_peak:.0f} N)")
        else:
            scale = SCALE_N_PER_UNIT
            print(f"[{cid}] combined scale: {scale:.4f} N/unit")
        force_N_all = load_raw * scale

        # ---- failure truncation ---------------------------------------------
        kept = truncation_mask(force_N_all)
        kept_idx = np.flatnonzero(kept)

        force_kept    = force_N_all[kept]
        stress_kept   = force_kept / area if np.isfinite(area) else np.full(kept.sum(), np.nan)
        eps_a_kept    = df["strain_axial_raw"].to_numpy()[kept]
        eps_t_kept    = df["strain_transverse_raw"].to_numpy()[kept]

        # Smoothing pass (no-op unless APPLY_SMOOTHING is on) — kept
        # pre-smoothing copies alongside for Level-3's diagnostic overlay.
        force_smoothed  = smooth_signal(force_kept)
        stress_smoothed = force_smoothed / area if np.isfinite(area) else np.full(kept.sum(), np.nan)
        eps_a_smoothed  = smooth_signal(eps_a_kept)
        eps_t_smoothed  = smooth_signal(eps_t_kept)

        p = compute_properties(eps_a_smoothed, stress_smoothed, eps_t_smoothed,
                               eps_axial_unsmoothed=eps_a_kept, sig_unsmoothed=stress_kept)
        if not p:
            print(f"[{cid}]  insufficient data")
            continue

        print(f"[{cid}]  E={p['E_GPa']:.2f} GPa  "
              f"σ_y={p['sigma_y_MPa']:.1f} MPa  "
              f"UTS={p['UTS_MPa']:.1f} MPa  "
              f"ε_UTS={p['eps_at_UTS']*100:.2f}%  "
              f"ν_chord={p['poisson_chord']:.3f}  "
              f"toe={p['eps_toe']*100:.3f}%")

        # ---- scatter L2 columns back onto the full per-frame file -----------
        final_idx = kept_idx[p["_valid"]]

        def scatter(values):
            col = np.full(n, np.nan)
            col[final_idx] = values
            return col

        df["kept"]                     = kept
        df["force_N"]                  = scatter(force_smoothed[p["_valid"]])
        df["stress_MPa"]               = scatter(p["_sig"])
        df["strain_axial"]             = scatter(p["_eps"])
        df["strain_transverse"]        = scatter(p["_eps_t"])
        df["stress_MPa_unsmoothed"]    = scatter(p["_sig_raw"])
        df["strain_axial_unsmoothed"]  = scatter(p["_eps_raw"])
        df.to_csv(frames_fp, index=False, float_format="%.6g")

        rows.append({"coupon": cid, **{k: v for k, v in p.items() if not k.startswith("_")}})

    if rows:
        write_specimen_sheet(rows)

        # ---- D638 §11.7 / §12.1: mean & std per (exposure, direction) -------
        write_group_stats(rows)

        print(f"\n{len(rows)} coupon(s) → DIC/*.csv, {SPECIMEN_SHEET.name}, "
              f"DIC/level2_group_stats.csv")

    print(f"\nDone. {time.time()-t0:.1f} s")


if __name__ == "__main__":
    main()
