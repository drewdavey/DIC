#!/usr/bin/env python3
"""
DIC_Level2.py  —  FSR Tensile Coupons
======================================
Reads Level-1 per-coupon CSVs, scales raw load to force/stress, applies
failure truncation and a light smoothing pass (median or Butterworth, see
FILTER_METHOD), and computes ASTM D638 mechanical properties. No plotting
here — see DIC_Level3.py.

Standards compliance — what each calculation cites
  Toe compensation     : D638 Annex A1 (mandatory unless toe is real material response)
  Modulus              : D638 §11.4   (slope of initial linear region of σ-ε)
  0.2 % offset yield    : D638 §A2.6 / Fig. A2.1 (offset from toe-corrected origin)
  Tensile strength UTS  : D638 §11.2   (max stress / original area)
  Poisson's ratio       : D638 Annex A3.10.1.3 (chord at ε_a=0.002 over 0.0005-0.0025)
  Group statistics      : D638 §11.7 / §12.1   (mean, std per series)

PROCESSING NOTE
  Level-1 writes the full, untruncated record (see its docstring). All
  failure truncation happens here via truncate_df(): pre-load slack and
  post-fracture rebound (past 50% post-UTS load drop) are cut. Properties
  are computed from this truncated, smoothed window; pre-smoothing values
  (still truncated) are kept alongside for Level-3's diagnostic overlay.

OUTPUT per coupon
  <DIC_DIR>/<coupon_id>_L2.csv   step, eps, sig, eps_t, i_uts, eps_raw, sig_raw
                                 (eps/sig/eps_t are toe-corrected + smoothed;
                                  eps_raw/sig_raw are the same window pre-smoothing)
  FSR-SpecimenTesting.xlsx       scalar properties written into each coupon's
                                 row (E, toe strain, yield stress/strain, UTS,
                                 strain at UTS, Poisson's ratio) — the single
                                 source of truth for per-coupon scalars used
                                 by Level-3 and the group-plot scripts.
  <DIC_DIR>/level2_group_stats.csv   D638 §11.7 mean/std/count by exposure×direction
"""

from __future__ import annotations
import sys
import time
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.ndimage import median_filter
from scipy.signal import butter, filtfilt
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# =============================================================================
# PATHS
# =============================================================================
DIC_DIR = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\04_TestCoupons"
    r"\P01-LT150-LH4.5\DIC"
)
SPECIMEN_SHEET = Path(
    r"Z:\2023_07_SIO_Functional_Surfing_Reef\04_Drew"
    r"\01_MaterialTesting\02_Mechanical Testing\FSR-SpecimenTesting.xlsx"
)

# =============================================================================
# SWITCHES
# =============================================================================
PRINTS     = ["P01"]
EXPOSURES  = {"CL": True, "UV": True, "SW": True, "IS": True}
DIRECTIONS = {"00": True, "45": True, "90": True}
REPLICATES = ["01", "02", "03"]

APPLY_SMOOTHING = True   # toggle the smoothing pass below on/off
FILTER_METHOD   = "butterworth"  # "median" or "butterworth" — see SMOOTHING section below

# =============================================================================
# FAILURE TRUNCATION  — applied to Level-1 data before property extraction
# Trim pre-load slack and post-fracture rebound so only the valid test window
# is passed to compute_properties.
# =============================================================================
LOAD_START_FRAC = 0.02     # pre-load: drop frames before load exceeds this × peak
LOAD_END_FRAC   = 0.50     # post-fracture: cut first post-UTS frame where load < this × peak

# Scale factor applied to load_raw to produce force_N (N per sync-CSV unit).
# Two modes — the first one found is used:
#   1. Per-coupon  : mts_peak_N / max(raw load_raw)  — most accurate, always
#                    available since Level-1 always writes mts_peak_N
#   2. Combined    : SCALE_N_PER_UNIT below          — fallback safety net
SCALE_N_PER_UNIT: float = 555.5928

# =============================================================================
# PROPERTY SETTINGS
# =============================================================================
# Modulus fit window (axial strain, dimensionless).
# D638 §11.4: "initial linear portion of the load-extension curve".
# A window of 0.05–0.3% covers the typical linear region for stiff polymers
# and composites without including the toe. Adjust if the fit line on the
# generated plot doesn't sit on the linear segment.
MODULUS_STRAIN_RANGE = (0.0005, 0.003)

# D638 §A2.6 — 0.2% offset yield strength
YIELD_OFFSET = 0.002

# D638 §A3.10.1.3 — Poisson chord method window (when no clear proportionality)
# Chord computed at ε_a = 0.002 over the range 0.0005 to 0.0025 strain.
POISSON_RANGE = (0.0005, 0.0025)
POISSON_CHORD_AT = 0.002

# Scalar property columns written into SPECIMEN_SHEET, keyed by coupon
# ("Specimen ID") — maps the property dict key to the Excel column header.
# Level-3 and the group-plot scripts read these same headers back out.
SPECIMEN_SHEET_COLUMNS = {
    "E_GPa":         "E (GPa)",
    "eps_toe":       "Toe Strain",
    "sigma_y_MPa":   "Yield Stress (MPa)",
    "eps_y":         "Yield Strain",
    "UTS_MPa":       "UTS (MPa)",
    "eps_at_UTS":    "Strain at UTS",
    "poisson_chord": "Poisson's Ratio (chord)",
    "poisson_slope": "Poisson's Ratio (slope)",
}

# =============================================================================
# SMOOTHING  — FILTER_METHOD selects which of these is used
#   "median"     : rolling median. Window must be odd. Raise MEDIAN_WINDOW
#                  for noisier data. Preferred default — doesn't ring or
#                  systematically undershoot a sharp peak the way an
#                  averaging-based filter like Butterworth can.
#   "butterworth": zero-phase low-pass (filtfilt). Lower BUTTER_CUTOFF for
#                  heavier smoothing; raise BUTTER_ORDER for a sharper
#                  rolloff. Output is clipped to the raw data's range to
#                  suppress filtfilt ringing at the truncation edges.
# =============================================================================
MEDIAN_WINDOW = 31  # frames

BUTTER_ORDER  = 2     # filter order
BUTTER_CUTOFF = 0.1  # cutoff frequency, fraction of Nyquist (0-1)

# =============================================================================
# HELPERS
# =============================================================================
def coupon_id(p, e, d, r): return f"{p}-T{e}{d}-{r}"

def selected_coupons():
    return [coupon_id(p, e, d, r)
            for p in PRINTS
            for e, on in EXPOSURES.items() if on
            for d, on2 in DIRECTIONS.items() if on2
            for r in REPLICATES]

def parse_id(cid):
    """Return (exposure, direction_str) e.g. ('CL', '00')."""
    part = cid.split("-")[1]
    return part[1:-2], part[-2:]

def find_l1(cid):
    p = DIC_DIR / f"{cid}_L1.csv"
    return p if p.exists() else None

def smooth_signal(x):
    """Dispatches to the filter selected by FILTER_METHOD. No-op when
    APPLY_SMOOTHING is False."""
    x = np.asarray(x, dtype=float)
    if not APPLY_SMOOTHING:
        return x.copy()
    if FILTER_METHOD == "butterworth":
        return _smooth_butterworth(x)
    return _smooth_median(x)

def _smooth_median(x):
    """Rolling median. mode='nearest' avoids the zero-padding edge artifacts
    scipy.signal.medfilt has."""
    win = MEDIAN_WINDOW
    if win % 2 == 0:
        win -= 1
    if win < 1 or len(x) < win:
        return x.copy()
    return median_filter(x, size=win, mode="constant", cval=np.nan)

def _smooth_butterworth(x):
    """Zero-phase low-pass (filtfilt). NaN gaps are bridged by linear
    interpolation before filtering, then the original NaN positions are
    restored. Output is clipped to the raw data's range — filtfilt can ring
    past it, especially near the truncation edges."""
    n = len(x)
    padlen = 3 * BUTTER_ORDER
    nan_mask = ~np.isfinite(x)
    if nan_mask.all() or n <= padlen:
        return x.copy()
    xi = x.copy()
    if nan_mask.any():
        idx = np.arange(n)
        xi[nan_mask] = np.interp(idx[nan_mask], idx[~nan_mask], x[~nan_mask])
    b, a = butter(BUTTER_ORDER, BUTTER_CUTOFF, btype="low", analog=False)
    out = filtfilt(b, a, xi)
    out = np.clip(out, np.nanmin(xi), np.nanmax(xi))
    out[nan_mask] = np.nan
    return out

def write_specimen_sheet(rows: list[dict]) -> None:
    """Write each coupon's scalar properties into its row in SPECIMEN_SHEET,
    matched by Specimen ID. Adds any missing property columns at the end;
    everything else in the workbook (other rows, formulas, formatting) is
    left untouched. Skipped (with a warning) if the file can't be opened —
    e.g. if it's currently open in Excel.
    """
    try:
        wb = openpyxl.load_workbook(SPECIMEN_SHEET)
    except FileNotFoundError:
        print(f"[!] {SPECIMEN_SHEET} not found — skipping specimen sheet update")
        return
    ws = wb.active

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    id_col = header.get("Specimen ID")
    if id_col is None:
        print("[!] 'Specimen ID' column not found in specimen sheet — skipping update")
        return

    next_col = ws.max_column + 1
    for label in SPECIMEN_SHEET_COLUMNS.values():
        if label not in header:
            ws.cell(row=1, column=next_col, value=label)
            header[label] = next_col
            next_col += 1

    row_by_id = {ws.cell(row=r, column=id_col).value: r
                 for r in range(2, ws.max_row + 1)}

    for row in rows:
        r = row_by_id.get(row["coupon"])
        if r is None:
            continue
        for key, label in SPECIMEN_SHEET_COLUMNS.items():
            v = row.get(key)
            v = None if (v is None or not np.isfinite(v)) else v
            ws.cell(row=r, column=header[label], value=v)

    # openpyxl doesn't evaluate formulas, so re-saving drops the cached
    # values of every formula cell in the workbook (e.g. Width/Dia,
    # Computed Area) until something recalculates them. Force a full
    # recalculation on next open so they never appear blank.
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(SPECIMEN_SHEET)
    except PermissionError:
        print(f"[!] {SPECIMEN_SHEET} is open elsewhere — could not save properties to it")

def truncate_df(df: pd.DataFrame) -> pd.DataFrame:
    """Remove pre-load slack and post-fracture frames (Level-2 failure truncation)."""
    f = df["force_N"].to_numpy()
    peak = float(np.nanmax(np.abs(f)))
    if peak <= 0:
        return df
    i_uts = int(np.nanargmax(np.abs(f)))
    starts = np.where(np.abs(f) > LOAD_START_FRAC * peak)[0]
    i0 = int(starts[0]) if len(starts) else 0
    post = np.where(np.abs(f[i_uts:]) < LOAD_END_FRAC * peak)[0]
    i1 = int(i_uts + post[0]) if len(post) else len(f) - 1
    return df.iloc[i0:i1 + 1].reset_index(drop=True)


# =============================================================================
# COMPUTE PROPERTIES
# =============================================================================
def compute_properties(df):
    """
    D638-compliant property extraction.

    Returns a dict with E_GPa, eps_toe (toe-correction offset applied),
    sigma_y_MPa, eps_y, UTS_MPa, eps_at_UTS, poisson_chord, poisson_slope,
    plus i_uts (index of UTS in the original arrays). Also returns the
    toe-corrected eps/sig/eps_t arrays (smoothed) and their pre-smoothing
    counterparts (eps_raw/sig_raw) for Level-3's diagnostic overlay.
    """
    df = df.dropna(subset=["strain_axial", "stress_MPa"]).reset_index(drop=True)
    if len(df) < 10:
        return None

    eps_raw   = df["strain_axial"].to_numpy()
    sig       = df["stress_MPa"].to_numpy()
    eps_t_raw = (df["strain_transverse"].to_numpy()
                 if "strain_transverse" in df.columns
                 else np.full_like(eps_raw, np.nan))
    # Pre-smoothing reference, carried through only for Level-3's diagnostic
    # overlay — not used in any property calculation.
    eps_unsmoothed = df["strain_axial_raw"].to_numpy() if "strain_axial_raw" in df.columns else None
    sig_unsmoothed = df["stress_MPa_raw"].to_numpy()   if "stress_MPa_raw"   in df.columns else None

    # ---- 1. Modulus (D638 §11.4) --------------------------------------------
    lo, hi = MODULUS_STRAIN_RANGE
    mfit = (eps_raw >= lo) & (eps_raw <= hi) & np.isfinite(eps_raw) & np.isfinite(sig)
    if mfit.sum() < 3:
        return None
    slope, intercept = np.polyfit(eps_raw[mfit], sig[mfit], 1)
    E_MPa = float(slope)

    # ---- 2. Toe compensation (D638 Annex A1) --------------------------------
    # The fitted line σ = E·ε + b is extended back to σ = 0; that strain
    # (b/(-E)) is the "toe offset" — all strains are then measured from the
    # corrected origin. ε_corrected = ε_raw − ε_offset.
    eps_offset = -intercept / E_MPa if E_MPa != 0 else 0.0
    eps   = eps_raw   - eps_offset
    eps_unsmoothed_corr = (eps_unsmoothed - eps_offset) if eps_unsmoothed is not None else None
    # Transverse strain: subtract its value at the corrected zero of axial strain.
    # Find the index where corrected axial ≈ 0 and subtract that ε_t.
    if np.any(np.isfinite(eps_t_raw)):
        i0 = int(np.nanargmin(np.abs(eps)))
        eps_t = eps_t_raw - eps_t_raw[i0]
    else:
        eps_t = eps_t_raw.copy()

    # ---- 3. UTS (D638 §11.2 — max stress) -----------------------------------
    i_uts   = int(np.nanargmax(sig))
    uts     = float(sig[i_uts])
    eps_ult = float(eps[i_uts])

    # ---- 4. 0.2% offset yield (D638 §A2.6, Fig. A2.1) -----------------------
    # First crossing of σ-ε curve with the line σ = E·(ε − YIELD_OFFSET).
    sigma_y, eps_y = np.nan, np.nan
    diff  = sig - E_MPa * (eps - YIELD_OFFSET)
    valid = np.where(eps > YIELD_OFFSET)[0]
    if len(valid) > 1:
        d = diff[valid]
        crossings = np.where(np.diff(np.sign(d)) < 0)[0]
        if len(crossings):
            k = valid[crossings[0]]
            denom = diff[k] - diff[k+1]
            f = diff[k] / denom if denom != 0 else 0.0
            eps_y   = float(eps[k] + f * (eps[k+1] - eps[k]))
            sigma_y = float(sig[k] + f * (sig[k+1] - sig[k]))

    # ---- 5. Poisson's ratio (D638 §A3.10.1.3, chord at ε_a = 0.002) --------
    #   ν = − ε_t(at ε_a = 0.002) / 0.002
    # Range 0.0005 – 0.0025 strain. Also report least-squares slope (§A3.10.1.1)
    # for transparency when proportionality holds.
    nu_chord = nu_slope = np.nan
    pm = ((eps >= POISSON_RANGE[0]) & (eps <= POISSON_RANGE[1]) &
          np.isfinite(eps) & np.isfinite(eps_t))
    if pm.sum() >= 3:
        nu_slope = float(-np.polyfit(eps[pm], eps_t[pm], 1)[0])
        order = np.argsort(eps[pm])
        ea, et = eps[pm][order], eps_t[pm][order]
        if ea[0] <= POISSON_CHORD_AT <= ea[-1]:
            nu_chord = float(-np.interp(POISSON_CHORD_AT, ea, et) / POISSON_CHORD_AT)

    return {
        "E_GPa":         E_MPa / 1000.0,
        "eps_toe":       eps_offset,
        "sigma_y_MPa":   sigma_y,
        "eps_y":         eps_y,
        "UTS_MPa":       uts,
        "eps_at_UTS":    eps_ult,
        "poisson_chord": nu_chord,
        "poisson_slope": nu_slope,
        "_eps":          eps,        # toe-corrected strain, used in property calcs
        "_sig":          sig,
        "_eps_t":        eps_t,
        "_i_uts":        i_uts,
        "_eps_raw":      eps_unsmoothed_corr,  # diagnostic overlay only
        "_sig_raw":      sig_unsmoothed,
    }


# =============================================================================
# MAIN
# =============================================================================
def main():
    t0 = time.time()
    print("=" * 70)
    print("DIC_Level2 — scale, truncate, smooth, compute D638 properties")
    print("=" * 70)
    rows = []

    for cid in selected_coupons():
        l1 = find_l1(cid)
        if l1 is None:
            print(f"[{cid}] no _L1.csv — run Level 1 first")
            continue
        df = pd.read_csv(l1)

        # ---- scale load_raw -> force_N -> stress_MPa --------------------
        raw_peak = float(np.nanmax(np.abs(df["load_raw"].to_numpy())))
        if "mts_peak_N" in df.columns and raw_peak > 0:
            mts_peak = float(df["mts_peak_N"].iloc[0])
            scale = mts_peak / raw_peak
            print(f"[{cid}] per-coupon scale: {scale:.4f} N/unit  (MTS {mts_peak:.0f} N)")
        else:
            scale = SCALE_N_PER_UNIT
            print(f"[{cid}] combined scale: {scale:.4f} N/unit")
        area = float(df["area_mm2"].iloc[0]) if "area_mm2" in df.columns else np.nan
        df["force_N"]    = df["load_raw"].to_numpy() * scale
        df["stress_MPa"] = df["force_N"] / area if np.isfinite(area) else np.nan

        df = truncate_df(df)

        # Keep pre-smoothing copies (still truncated) for Level-3's
        # diagnostic overlay.
        df["stress_MPa_raw"]   = df["stress_MPa"]
        df["strain_axial_raw"] = df["strain_axial"]

        df["force_N"]      = smooth_signal(df["force_N"].to_numpy())
        df["stress_MPa"]   = df["force_N"] / area if np.isfinite(area) else np.nan
        df["strain_axial"] = smooth_signal(df["strain_axial"].to_numpy())
        if "strain_transverse" in df.columns:
            df["strain_transverse"] = smooth_signal(df["strain_transverse"].to_numpy())

        p = compute_properties(df)
        if not p:
            print(f"[{cid}]  insufficient data")
            continue

        print(f"[{cid}]  E={p['E_GPa']:.2f} GPa  "
              f"σ_y={p['sigma_y_MPa']:.1f} MPa  "
              f"UTS={p['UTS_MPa']:.1f} MPa  "
              f"ε_UTS={p['eps_at_UTS']*100:.2f}%  "
              f"ν_chord={p['poisson_chord']:.3f}  "
              f"toe={p['eps_toe']*100:.3f}%")

        # ---- write per-frame curve CSV for Level-3 plotting --------------
        n_frames = len(p["_eps"])
        eps_raw_col = p["_eps_raw"] if p["_eps_raw"] is not None else np.full(n_frames, np.nan)
        sig_raw_col = p["_sig_raw"] if p["_sig_raw"] is not None else np.full(n_frames, np.nan)
        pd.DataFrame({
            "step":    np.arange(n_frames),
            "eps":     p["_eps"],
            "sig":     p["_sig"],
            "eps_t":   p["_eps_t"],
            "i_uts":   p["_i_uts"],
            "eps_raw": eps_raw_col,
            "sig_raw": sig_raw_col,
        }).to_csv(DIC_DIR / f"{cid}_L2.csv", index=False, float_format="%.6g")

        rows.append({"coupon": cid, **{k: v for k, v in p.items() if not k.startswith("_")}})

    if rows:
        write_specimen_sheet(rows)

        # ---- D638 §11.7 / §12.1: mean & std per (exposure, direction) -------
        df_sum = pd.DataFrame(rows)
        df_sum["exposure"]  = df_sum["coupon"].map(lambda c: parse_id(c)[0])
        df_sum["direction"] = df_sum["coupon"].map(lambda c: parse_id(c)[1])
        agg_cols = ["E_GPa", "sigma_y_MPa", "UTS_MPa", "eps_at_UTS", "poisson_chord"]
        group = (df_sum.groupby(["exposure", "direction"])[agg_cols]
                       .agg(["mean", "std", "count"]))
        group.to_csv(DIC_DIR / "level2_group_stats.csv")

        print(f"\n{len(rows)} coupon(s) → DIC/*_L2.csv, {SPECIMEN_SHEET.name}, "
              f"DIC/level2_group_stats.csv")

    print(f"\nDone. {time.time()-t0:.1f} s")


if __name__ == "__main__":
    main()
